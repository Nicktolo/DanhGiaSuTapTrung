from keras.models import load_model
from time import sleep, strftime, gmtime
from tensorflow.keras.utils import img_to_array
from keras.preprocessing import image
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment
import cv2
import numpy as np


face_classifier = cv2.CascadeClassifier(r'haarcascade_frontalface_default.xml')
classifier = load_model(r'model.h5')

emotion_labels = ['Angry-Not focus', 'Disgust-Not focus', 'Fear-Not focus', 'Happy-Not focus', 'Neutral-focus', 'Sad-Not focus', 'Surprise-Not focus']

cap = cv2.VideoCapture(0)

emotion_count = {emotion: 0 for emotion in emotion_labels}
emotion_history = []  
total_frames = 0
last_emotion = None  

while True:
    _, frame = cap.read()
    gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
    faces = face_classifier.detectMultiScale(gray)

    if len(faces) == 1:  
        (x, y, w, h) = faces[0]
        cv2.rectangle(frame, (x, y), (x+w, y+h), (0, 255, 255), 2)
        roi_gray = gray[y:y+h, x:x+w]
        roi_gray = cv2.resize(roi_gray, (48, 48), interpolation=cv2.INTER_AREA)

        if np.sum([roi_gray]) != 0:
            roi = roi_gray.astype('float')/255.0
            roi = img_to_array(roi)
            roi = np.expand_dims(roi, axis=0)

            prediction = classifier.predict(roi)[0]
            label = emotion_labels[prediction.argmax()]
            emotion_count[label] += 1
            total_frames += 1

            
            if label != last_emotion:
                last_emotion = label
                timestamp = strftime("%d-%m-%Y %H:%M:%S", gmtime())  
                emotion_history.append((label, timestamp))
                print(f"New emotion detected: {label} at {timestamp}")

            label_position = (x, y)
            cv2.putText(frame, label, label_position, cv2.FONT_HERSHEY_SIMPLEX, 1, (0, 255, 0), 2)
        else:
            cv2.putText(frame, 'No Faces', (30, 80), cv2.FONT_HERSHEY_SIMPLEX, 1, (0, 255, 0), 2)
    elif len(faces) > 1:
        cv2.putText(frame, 'Multiple Faces Detected', (30, 80), cv2.FONT_HERSHEY_SIMPLEX, 1, (0, 0, 255), 2)
    
    cv2.imshow('Emotion Detector', frame)
    if cv2.waitKey(1) & 0xFF == ord('q'):
        break

cap.release()
cv2.destroyAllWindows()


if total_frames > 0:
    emotion_percentages = {emotion: (count / total_frames) * 100 for emotion, count in emotion_count.items()}
    print("\nEmotion counts and percentages:")
    for emotion, count in emotion_count.items():
        percentage = emotion_percentages[emotion]
        print(f"{emotion}: {count} times ({percentage:.2f}%)")

    most_common_emotion = max(emotion_count, key=emotion_count.get)

    print(f"\nKết quả dự đoán cảm xúc người học: {most_common_emotion}({emotion_percentages[most_common_emotion]:.2f}%)")
    print("\nEmotion history:")
    for emotion, timestamp in emotion_history:
        print(f"{emotion} detected at {timestamp}")
    
    try:
        wb = load_workbook('History.xlsx')
    except:
        wb = Workbook()
    timestamp = strftime("%d-%m-%Y_%H-%M-%S") 
    sheet_name = f"{timestamp}"
    ws = wb.create_sheet(title=sheet_name)

    history_text = "\n".join([f"{emotion}: {timestamp}" for timestamp, emotion in emotion_history])
    
    ws['A1'] = "Emotion History"
    ws['A2'] = history_text
    ws['A2'].alignment = Alignment(wrap_text=True) 

    wb.save('History.xlsx')
else:
    print("No emotions detected.")
